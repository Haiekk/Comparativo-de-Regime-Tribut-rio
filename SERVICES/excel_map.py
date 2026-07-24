CAMPOS_TEXTO = {"razao_social", "regime_atual", "atividade"}

INPUTS = {

    "empresa": {
        "razao_social": ("Dados do Cliente", "B5"),
        "regime_atual": ("Dados do Cliente", "B6"),
        "atividade": ("Dados do Cliente", "B7"),
    },

    "receitas": {
        "receita_comercio": ("Dados do Cliente", "B10"),
        "receita_servicos": ("Dados do Cliente", "B11"),
        "receita_sem_nota": ("Dados do Cliente", "B12"),
        "rbt12": ("Dados do Cliente", "B14"),
    },

    "creditos": {
        "mercadorias_revenda": ("Dados do Cliente", "B18"),
        "aluguel_pj": ("Dados do Cliente", "B19"),
        "frete_compras": ("Dados do Cliente", "B20"),
        "frete_vendas": ("Dados do Cliente", "B21"),
        "servicos_terceiros": ("Dados do Cliente", "B22"),
        "combustiveis": ("Dados do Cliente", "B23"),
        "epis": ("Dados do Cliente", "B24"),
        "energia": ("Dados do Cliente", "B25"),
        "depreciacao": ("Dados do Cliente", "B26"),
        "ferramentas": ("Dados do Cliente", "B27"),
        "outras_despesas": ("Dados do Cliente", "B28"),
    },

    "despesas": {
        "compras": ("Dados do Cliente", "B32"),
        "despesas_operacionais": ("Dados do Cliente", "B33"),
        "despesas_aluguel": ("Dados do Cliente", "B34"),
        "materiais_consumo": ("Dados do Cliente", "B35"),
        "materiais_limpeza": ("Dados do Cliente", "B36"),
        "despesas_energia": ("Dados do Cliente", "B37"),
        "despesas_financeiras": ("Dados do Cliente", "B38"),
        "despesas_combustivel": ("Dados do Cliente", "B39"),
    },

    "folha": {
        "folha_pagamento": ("Dados do Cliente", "B45"),
        "ticket_alimentacao": ("Dados do Cliente", "B46"),
    },

}

CAMPOS_ESPELHADOS = {}


OUTPUTS = {

    "comparativo": {
        "simples_total_impostos": ("Comparativo de Regimes", "B10"),
        "presumido_total_impostos": ("Comparativo de Regimes", "C10"),
        "real_total_impostos": ("Comparativo de Regimes", "D10"),

        "simples_carga": ("Comparativo de Regimes", "B11"),
        "presumido_carga": ("Comparativo de Regimes", "C11"),
        "real_carga": ("Comparativo de Regimes", "D11"),

        "simples_folha": ("Comparativo de Regimes", "B14"),
        "presumido_folha": ("Comparativo de Regimes", "C14"),
        "real_folha": ("Comparativo de Regimes", "D14"),

        "simples_custo_total": ("Comparativo de Regimes", "B15"),
        "presumido_custo_total": ("Comparativo de Regimes", "C15"),
        "real_custo_total": ("Comparativo de Regimes", "D15"),

        "simples_lucro": ("Comparativo de Regimes", "B16"),
        "presumido_lucro": ("Comparativo de Regimes", "C16"),
        "real_lucro": ("Comparativo de Regimes", "D16"),

        "simples_recomendado": ("Comparativo de Regimes", "B18"),
        "presumido_recomendado": ("Comparativo de Regimes", "C18"),
        "real_recomendado": ("Comparativo de Regimes", "D18"),
    },

    "detalhamento": {
        "simples_das_irpj": ("Comparativo de Regimes", "B5"),
        "presumido_das_irpj": ("Comparativo de Regimes", "C5"),
        "real_das_irpj": ("Comparativo de Regimes", "D5"),

        "simples_csll": ("Comparativo de Regimes", "B6"),
        "presumido_csll": ("Comparativo de Regimes", "C6"),
        "real_csll": ("Comparativo de Regimes", "D6"),

        "simples_pis": ("Comparativo de Regimes", "B7"),
        "presumido_pis": ("Comparativo de Regimes", "C7"),
        "real_pis": ("Comparativo de Regimes", "D7"),

        "simples_cofins": ("Comparativo de Regimes", "B8"),
        "presumido_cofins": ("Comparativo de Regimes", "C8"),
        "real_cofins": ("Comparativo de Regimes", "D8"),

        "simples_iss_icms": ("Comparativo de Regimes", "B9"),
        "presumido_iss_icms": ("Comparativo de Regimes", "C9"),
        "real_iss_icms": ("Comparativo de Regimes", "D9"),
    },

    "memoria_simples": {
        "aliquota_efetiva_anexo_i": ("Simples Nacional", "B8"),
        "das_anexo_i": ("Simples Nacional", "B10"),
        "aliquota_efetiva_anexo_iii": ("Simples Nacional", "B16"),
        "das_anexo_iii": ("Simples Nacional", "B18"),
    },

    "creditos": {
        "credito_pis": ("Dados do Cliente", "C29"),
        "credito_cofins": ("Dados do Cliente", "D29"),
        "base_lucro_real": ("Lucro Real", "B27"),
    },

    "dre": {
        "receita_bruta_mensal": ("Dados do Cliente", "B13"),

        "simples_receita_bruta": ("DRE Comparativa", "B6"),
        "presumido_receita_bruta": ("DRE Comparativa", "C6"),
        "real_receita_bruta": ("DRE Comparativa", "D6"),

        "simples_das": ("DRE Comparativa", "B7"),
        "presumido_das": ("DRE Comparativa", "C7"),
        "real_das": ("DRE Comparativa", "D7"),

        "simples_icms": ("DRE Comparativa", "B8"),
        "presumido_icms": ("DRE Comparativa", "C8"),
        "real_icms": ("DRE Comparativa", "D8"),

        "simples_iss": ("DRE Comparativa", "B9"),
        "presumido_iss": ("DRE Comparativa", "C9"),
        "real_iss": ("DRE Comparativa", "D9"),

        "simples_pis": ("DRE Comparativa", "B10"),
        "presumido_pis": ("DRE Comparativa", "C10"),
        "real_pis": ("DRE Comparativa", "D10"),

        "simples_cofins": ("DRE Comparativa", "B11"),
        "presumido_cofins": ("DRE Comparativa", "C11"),
        "real_cofins": ("DRE Comparativa", "D11"),

        "simples_receita_liquida": ("DRE Comparativa", "B12"),
        "presumido_receita_liquida": ("DRE Comparativa", "C12"),
        "real_receita_liquida": ("DRE Comparativa", "D12"),

        "simples_cmv": ("DRE Comparativa", "B13"),
        "presumido_cmv": ("DRE Comparativa", "C13"),
        "real_cmv": ("DRE Comparativa", "D13"),

        "simples_lucro_bruto": ("DRE Comparativa", "B14"),
        "presumido_lucro_bruto": ("DRE Comparativa", "C14"),
        "real_lucro_bruto": ("DRE Comparativa", "D14"),

        "simples_pessoal": ("DRE Comparativa", "B15"),
        "presumido_pessoal": ("DRE Comparativa", "C15"),
        "real_pessoal": ("DRE Comparativa", "D15"),

        "simples_aluguel": ("DRE Comparativa", "B16"),
        "presumido_aluguel": ("DRE Comparativa", "C16"),
        "real_aluguel": ("DRE Comparativa", "D16"),

        "simples_consumo": ("DRE Comparativa", "B17"),
        "presumido_consumo": ("DRE Comparativa", "C17"),
        "real_consumo": ("DRE Comparativa", "D17"),

        "simples_limpeza": ("DRE Comparativa", "B18"),
        "presumido_limpeza": ("DRE Comparativa", "C18"),
        "real_limpeza": ("DRE Comparativa", "D18"),

        "simples_energia": ("DRE Comparativa", "B19"),
        "presumido_energia": ("DRE Comparativa", "C19"),
        "real_energia": ("DRE Comparativa", "D19"),

        "simples_financeiras": ("DRE Comparativa", "B20"),
        "presumido_financeiras": ("DRE Comparativa", "C20"),
        "real_financeiras": ("DRE Comparativa", "D20"),

        "simples_combustivel": ("DRE Comparativa", "B21"),
        "presumido_combustivel": ("DRE Comparativa", "C21"),
        "real_combustivel": ("DRE Comparativa", "D21"),

        "simples_alimentacao": ("DRE Comparativa", "B22"),
        "presumido_alimentacao": ("DRE Comparativa", "C22"),
        "real_alimentacao": ("DRE Comparativa", "D22"),

        "simples_outras": ("DRE Comparativa", "B23"),
        "presumido_outras": ("DRE Comparativa", "C23"),
        "real_outras": ("DRE Comparativa", "D23"),

        "simples_resultado_antes": ("DRE Comparativa", "B24"),
        "presumido_resultado_antes": ("DRE Comparativa", "C24"),
        "real_resultado_antes": ("DRE Comparativa", "D24"),

        "simples_irpj": ("DRE Comparativa", "B25"),
        "presumido_irpj": ("DRE Comparativa", "C25"),
        "real_irpj": ("DRE Comparativa", "D25"),

        "simples_csll": ("DRE Comparativa", "B26"),
        "presumido_csll": ("DRE Comparativa", "C26"),
        "real_csll": ("DRE Comparativa", "D26"),

        "simples_lucro_liquido": ("DRE Comparativa", "B27"),
        "presumido_lucro_liquido": ("DRE Comparativa", "C27"),
        "real_lucro_liquido": ("DRE Comparativa", "D27"),

        "simples_margem": ("DRE Comparativa", "B29"),
        "presumido_margem": ("DRE Comparativa", "C29"),
        "real_margem": ("DRE Comparativa", "D29"),
    },


}

ANCORAS = {
    ("Comparativo de Regimes", 5): "DAS / IRPJ",
    ("Comparativo de Regimes", 6): "CSLL",
    ("Comparativo de Regimes", 7): "PIS",
    ("Comparativo de Regimes", 8): "COFINS",
    ("Comparativo de Regimes", 9): "ISS + ICMS",
    ("Comparativo de Regimes", 10): "TOTAL DE IMPOSTOS",
    ("Comparativo de Regimes", 11): "CARGA",
    ("Comparativo de Regimes", 14): "CUSTO TOTAL DE FOLHA",
    ("Comparativo de Regimes", 15): "CUSTO TOTAL",
    ("Comparativo de Regimes", 16): "LUCRO LÍQUIDO",
    ("Comparativo de Regimes", 18): "REGIME RECOMENDADO",

    ("DRE Comparativa", 8): "ICMS SOBRE VENDAS",
    ("DRE Comparativa", 27): "LUCRO LÍQUIDO DO PERÍODO",
    ("DRE Comparativa", 29): "MARGEM LÍQUIDA",

    ("Simples Nacional", 8): "ALÍQUOTA EFETIVA",
    ("Simples Nacional", 10): "DAS MENSAL",
    ("Simples Nacional", 16): "ALÍQUOTA EFETIVA",
    ("Simples Nacional", 18): "DAS MENSAL",

    ("Dados do Cliente", 13): "FATURAMENTO MENSAL TOTAL",
    ("Dados do Cliente", 29): "TOTAL DE CRÉDITOS",

    ("Lucro Real", 28): "LUCRO REAL",
}


def _celulas_mapeadas():
    for grupo in INPUTS.values():
        for campo, alvo in grupo.items():
            yield "entrada", campo, alvo
    for campo, alvos in CAMPOS_ESPELHADOS.items():
        for alvo in alvos:
            yield "entrada", campo, alvo
    for grupo in OUTPUTS.values():
        for campo, alvo in grupo.items():
            yield "saida", campo, alvo


def validar(caminho_modelo):
    import re

    from openpyxl import load_workbook

    wb = load_workbook(caminho_modelo)
    problemas = []
    abas = set(wb.sheetnames)

    for papel, campo, (aba, celula) in _celulas_mapeadas():
        if aba not in abas:
            problemas.append(f"{campo}: aba '{aba}' não existe")
            continue

        valor = wb[aba][celula].value
        eh_formula = isinstance(valor, str) and valor.startswith("=")

        if papel == "saida" and not eh_formula:
            if isinstance(valor, (int, float)):
                continue        
            problemas.append(
                f"{campo} ({aba}!{celula}): esperava fórmula, encontrou {valor!r} — "
                "linha deslocada ou fórmula sobrescrita por valor fixo"
            )
        elif papel == "entrada" and eh_formula:
            problemas.append(
                f"{campo} ({aba}!{celula}): é uma fórmula ({valor!r}) e seria "
                "sobrescrita ao preencher"
            )

    for (aba, linha), esperado in ANCORAS.items():
        if aba not in abas:
            problemas.append(f"âncora: aba '{aba}' não existe")
            continue
        rotulo = str(wb[aba][f"A{linha}"].value or "").upper()
        if esperado.upper() not in rotulo:
            problemas.append(
                f"âncora {aba}!A{linha}: esperava conter {esperado!r}, "
                f"encontrou {rotulo[:45]!r} — linhas deslocadas"
            )

    return problemas


if __name__ == "__main__":
    import sys
    from pathlib import Path

    modelo = Path(sys.argv[1]) if len(sys.argv) > 1 else (
        Path(__file__).resolve().parent.parent / "EXCEL" / "Greentax_Planejamento_Tributario_Compressores(Recuperado Automaticamente).xlsx"
    )
    falhas = validar(modelo)
    if falhas:
        print(f"{len(falhas)} problema(s) no mapa:")
        for f in falhas:
            print("  -", f)
        sys.exit(1)
    total = sum(1 for _ in _celulas_mapeadas())
    print(f"Mapa consistente com {modelo.name}: {total} células verificadas.")